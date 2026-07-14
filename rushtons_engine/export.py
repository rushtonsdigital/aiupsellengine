"""Excel tracker: the human review surface for Phase 1.

Summary tab + one tab per recommended account. The DB stays authoritative;
this file is an export of `recommendations` + `comms`, with blank columns the
CS team fills in by hand (contact, approved, sent, outcome).
"""

import logging
import re
import time
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

import config

log = logging.getLogger(__name__)

HEADER_FILL = PatternFill("solid", fgColor="1F6F43")   # Rushton's green
HEADER_FONT = Font(color="FFFFFF", bold=True)
LABEL_FONT = Font(bold=True)
WRAP = Alignment(wrap_text=True, vertical="top")

SUMMARY_COLS = [
    "Rank", "Customer Code", "Customer Name", "Venue Type", "Account Manager",
    "Size Band", "Orders (month)", "SKUs", "Categories Bought",
    "Gap Categories to Pitch", "Products Pitched", "Score",
    "WhatsApp - Announcement", "WhatsApp - Follow-up", "WhatsApp - Post-box",
    "Contact (phone)", "Approved (Y/N)", "Sent Date", "Box Sent (Y/N)",
    "Outcome", "Notes",
]
SUMMARY_COL_WIDTHS = [6, 14, 30, 16, 18, 10, 12, 8, 28, 30, 34, 8, 45, 45, 45,
                     16, 12, 12, 12, 18, 30]
WHATSAPP_COL_LABELS = ("WhatsApp - Announcement", "WhatsApp - Follow-up", "WhatsApp - Post-box")


def _pitched(rec: dict) -> str:
    """The products the drafter actually chose — blank until drafts applied."""
    return ", ".join(f"{p['name']} ({p['code']})"
                     for p in rec.get("chosen_products") or [])


def _wrapped_row_height(texts: list[str], chars_per_line: int = 45,
                        pt_per_line: int = 15, min_pt: int = 15) -> int:
    """Estimate row height for wrapped text in a single-line-per-message cell."""
    lines = max((len(t) + chars_per_line - 1) // chars_per_line for t in texts) if texts else 1
    return max(min_pt, pt_per_line * max(lines, 1))


def _save_with_fallback(wb: Workbook, path: Path,
                        max_retries: int = 3, retry_delay: float = 1.5) -> Path:
    """Save the workbook, retrying briefly for a transient lock (e.g. a
    background sync client scanning the file — common once this folder is
    synced to SharePoint/OneDrive). If it's still locked after that — most
    likely a human has it open in Excel — write to a clearly-named fallback
    path instead of crashing the whole weekly run. Either way, the database
    is already the authoritative record by the time this runs; the tracker
    is just the human-facing export."""
    for attempt in range(max_retries):
        try:
            wb.save(path)
            return path
        except PermissionError:
            if attempt < max_retries - 1:
                log.warning("tracker file locked, retrying (%d/%d): %s",
                            attempt + 1, max_retries, path)
                time.sleep(retry_delay)
    fallback = path.with_stem(f"{path.stem}_UPDATED_{datetime.now():%H%M}")
    wb.save(fallback)
    log.error("%s is locked (likely open in Excel) - wrote %s instead. Close "
             "the original and replace it with this one, or just keep "
             "working from the new file.", path, fallback)
    return fallback


def _safe_tab_name(name: str, used: set) -> str:
    tab = re.sub(r"[\[\]:*?/\\]", "", name)[:28] or "Account"
    base, i = tab, 2
    while tab in used:
        tab = f"{base[:25]}_{i}"
        i += 1
    used.add(tab)
    return tab


def _header_row(ws, row, values):
    for col, value in enumerate(values, start=1):
        cell = ws.cell(row=row, column=col, value=value)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT


def write_tracker(recommendations: list[dict], run_date, out_dir: Path | None = None) -> Path:
    out_dir = Path(out_dir or config.EXPORT_DIR)
    out_dir.mkdir(parents=True, exist_ok=True)
    path = out_dir / f"rushtons_upsell_tracker_{run_date}.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    ws.cell(row=1, column=1, value=f"Rushton's Upsell Engine — week of {run_date}").font = Font(bold=True, size=14)
    _header_row(ws, 3, SUMMARY_COLS)
    for i, rec in enumerate(recommendations, start=4):
        drafts = rec.get("drafts", {})
        messages = {"WhatsApp - Announcement": drafts.get("announcement", ""),
                   "WhatsApp - Follow-up": drafts.get("followup", ""),
                   "WhatsApp - Post-box": drafts.get("postbox", "")}
        row_values = {
            "Rank": rec["rank"], "Customer Code": rec["customer_code"],
            "Customer Name": rec["customer_name"], "Venue Type": rec["venue_type"],
            "Account Manager": rec["sales_rep"], "Size Band": rec["size_band"],
            "Orders (month)": rec["num_orders"], "SKUs": rec["num_skus"],
            "Categories Bought": ", ".join(rec["bought_categories"]),
            "Gap Categories to Pitch": ", ".join(rec["gap_categories"]),
            "Products Pitched": _pitched(rec),
            "Score": float(rec["score"]),
            **messages,
            "Contact (phone)": "", "Approved (Y/N)": "", "Sent Date": "",
            "Box Sent (Y/N)": "", "Outcome": "", "Notes": "",
        }
        for col, label in enumerate(SUMMARY_COLS, start=1):
            cell = ws.cell(row=i, column=col, value=row_values[label])
            if label in WHATSAPP_COL_LABELS:
                cell.alignment = WRAP
        ws.row_dimensions[i].height = _wrapped_row_height(list(messages.values()))
    for col, width in enumerate(SUMMARY_COL_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.freeze_panes = "A4"

    used_tabs = {"Summary"}
    for rec in recommendations:
        tab = wb.create_sheet(_safe_tab_name(rec["customer_name"], used_tabs))
        tab.column_dimensions["A"].width = 22
        tab.column_dimensions["B"].width = 90
        row = 1

        def put(label, value, wrap=False):
            nonlocal row
            tab.cell(row=row, column=1, value=label).font = LABEL_FONT
            cell = tab.cell(row=row, column=2, value=value)
            if wrap:
                cell.alignment = WRAP
            row += 1

        put("Customer", f"{rec['customer_name']} ({rec['customer_code']})")
        put("Venue type", rec["venue_type"])
        put("Account manager", rec["sales_rep"])
        put("Rank / score", f"{rec['rank']} / {rec['score']}")
        put("Why selected", rec["rationale"], wrap=True)
        put("Currently buys", ", ".join(rec["bought_categories"]))
        if rec.get("customer_review"):
            put("Customer review", rec["customer_review"], wrap=True)
        if rec.get("data_notes"):
            put("Data notes", rec["data_notes"], wrap=True)
        row += 1

        # What we're actually pitching, and why — the review surface that matters.
        chosen = rec.get("chosen_products") or []
        _header_row(tab, row, ["Product pitched", "Why this kitchen"])
        row += 1
        for p in chosen:
            tab.cell(row=row, column=1,
                     value=f"{p['name']} ({p['code']})").font = LABEL_FONT
            cell = tab.cell(row=row, column=2, value=p.get("why", ""))
            cell.alignment = WRAP
            row += 1
        if not chosen:
            tab.cell(row=row, column=1, value="[pending]").font = LABEL_FONT
            tab.cell(row=row, column=2, value="products chosen at drafting")
            row += 1
        row += 1

        # The full eligible pool, for anyone who wants to check the call.
        _header_row(tab, row, ["Gap category", "Eligible pool (in season, not currently bought)"])
        row += 1
        for cat, items in rec["product_pool"].items():
            names = "\n".join(f"{p['name']} ({p['code']}) — {p['buyers_14d']} recent buyers"
                              for p in items) or "no in-season products available"
            tab.cell(row=row, column=1, value=cat).font = LABEL_FONT
            cell = tab.cell(row=row, column=2, value=names)
            cell.alignment = WRAP
            tab.row_dimensions[row].height = 15 * max(len(items), 1)
            row += 1
        row += 1
        _header_row(tab, row, ["Stage", "Draft (review before sending)"])
        row += 1
        stage_labels = {"announcement": "1. WhatsApp opener",
                        "followup": "2. Follow-up (if no reply)",
                        "postbox": "3. Post-box check-in"}
        for stage, label in stage_labels.items():
            body = rec.get("drafts", {}).get(stage, "")
            tab.cell(row=row, column=1, value=label).font = LABEL_FONT
            cell = tab.cell(row=row, column=2, value=body)
            cell.alignment = WRAP
            tab.row_dimensions[row].height = max(30, 14 * (body.count("\n") + 2))
            row += 1
        row += 1
        for label in ("Contact (phone)", "Approved by", "Sent date",
                      "Box sent (Y/N)", "Outcome", "Notes"):
            put(label, "")

    path = _save_with_fallback(wb, path)
    log.info("tracker written: %s (%d account tabs)", path, len(recommendations))
    return path
